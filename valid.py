from openpyxl import load_workbook

workbook_old = load_workbook(filename="电路分析2023春季.xlsx")
workbook_new = load_workbook(filename="new_save_sheet.xlsx")

sheet_old = workbook_old["工作表1"]
sheet_new = workbook_new["sheet1"]



## read old workbook and store all student name-grade as list
student_name_old = []
student_grade_old = []
for i in range(150):
    student_name_position  = 'C' + str(i + 2)
    student_grade_position = 'F' + str(i + 2)
    student_name = sheet_old[student_name_position].value
    student_grade = sheet_old[student_grade_position].value
    student_name_old.append(student_name)
    student_grade_old.append(student_grade)

student_name_new = []
student_grade_new = []
for i in range(160):
    student_name_position  = 'C' + str(i + 2)
    student_grade_position = 'F' + str(i + 2)
    student_name = sheet_new[student_name_position].value
    student_grade = sheet_new[student_grade_position].value
    student_name_new.append(student_name)
    student_grade_new.append(student_grade)
#
# for i in range(150):
#     print(i+1,student_name_old[i],student_grade_old[i])
#
# for i in range(160):
#     print(i+1,student_name_new[i],student_grade_new[i])

for old_name in student_name_old:
    if old_name in student_name_new:
        old_index = student_name_old.index(old_name)
        new_index = student_name_new.index(old_name)
        if(student_grade_old[old_index]!=student_grade_new[new_index]):
            print("not equal!")

for new_name in student_name_new:
    if new_name in student_name_old:
        old_index = student_name_old.index(new_name)
        new_index = student_name_new.index(new_name)
        if(student_grade_old[old_index]!=student_grade_new[new_index]):
            print("not equal!")

print("done without problem")