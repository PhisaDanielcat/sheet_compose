from openpyxl import load_workbook

workbook_old = load_workbook(filename="电路分析2023春季.xlsx")
workbook_new = load_workbook(filename="2023电路分析名单new.xlsx")

sheet_old = workbook_old["工作表1"]
sheet_new = workbook_new["sheet1"]



## read old workbook and store all student name-grade as list
student_name_old = []
student_grade_old = []
for i in range(150):
    student_name_position  = 'C' + str(i + 2)
    student_grade_position = 'F' + str(i + 2)
    student_name = sheet_old[student_name_position].value
    student_grade = sheet_old[student_grade_position].value
    student_name_old.append(student_name)
    student_grade_old.append(student_grade)

#
## read new workbook and store all student names as list
student_name_new = []
for i in range(160):
    student_name_position  = 'C' + str(i + 2)
    student_name = sheet_new[student_name_position].value
    # print(i+1,student_name)
    student_name_new.append(student_name)

#####################################################
## who drop out from class
# for name in student_name_old:
#     if name not in student_name_new:
#         print(name,"    not match!")

## who newly comes in to class
# for name in student_name_new:
#     if name not in student_name_old:
#         print(name,"    newly comes!")
######################################################

## match and create new grade list
i=0
student_grade_new = [None for _ in range(len(student_name_new))]
for old_name in student_name_old:
    for new_name in student_name_new:
        if(old_name == new_name):
            # i+=1
            # print(i,"match",old_name,"with",new_name,"in position",student_name_new.index(old_name))
            student_grade_new[student_name_new.index(old_name)] = student_grade_old[student_name_old.index(old_name)]

for i in range(len(student_name_new)):
    print(i,student_name_new[i],student_grade_new[i])

    new_grade_position = "F" + str(i+2)
    sheet_new[new_grade_position].value = student_grade_new[i]


workbook_new.save("new_save_sheet.xlsx")
